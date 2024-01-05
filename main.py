import openpyxl 
import pandas 
from datetime import datetime

#محصولات غذایی
class Food_Products:
    def __init__(self, file_name):
        self.file_name = file_name
        self.workbook = openpyxl.load_workbook(file_name)
        self.sheet = self.workbook.active

#نمایش دادن محصولات غذایی  
    def show_products(self):
        products_list=pandas.read_excel(self.file_name)
        print (products_list)

#اضافه کردن محصولات غذایی
    def add_product(self, product_code, product_name, company, date_of_manufacturing,expiry,bought,stock,price):
      
        self.sheet.append(["product Code", "product name", "company", "date_of_manufacturing", "expiry", "bought", "stock","price"])
        self.sheet.append([ product_code, product_name, company, date_of_manufacturing,expiry,bought,stock,price])
        self.workbook.save(self.file_name)
       
# تغییر دادن محصولات غذایی      
    def edit_product(self, row_num, new_info):
        for col, value in enumerate(new_info, start=1):
            self.sheet.cell(row=row_num+2, column=col, value=value)
        self.workbook.save(self.file_name)

#حذف کردن محصولات غذایی    
    def delete_product(self, row_num):
        self.sheet.delete_rows(row_num+2)
        self.workbook.save(self.file_name)
        

#مشتریان
class Customer:
    def __init__(self, cfile_name):
        self.cfile_name=cfile_name
        self.workbook=openpyxl.load_workbook(cfile_name)
        self.sheet=self.workbook.active
        
#نمایش دادن مشتریان  
    def show_customers(self):
        customers_list=pandas.read_excel(self.cfile_name)
        print (customers_list)
    
#اضافه کردن مشتریان
    def add_customers(self,username,name,family,personal_id,phone_number,purchase,discount,credit,debt):
        self.sheet.append(["username", "name", "family", "personal_id", "phone_number", "purchase", "discount","credit","debt"])
        self.sheet.append([ username,name,family,personal_id,phone_number,purchase,discount,credit,debt])
        self.workbook.save(self.cfile_name)
              
# تغییر دادن مشتریان      
    def edit_customers(self, row_num, new_info):
        for col, value in enumerate(new_info, start=1):
            self.sheet.cell(row=row_num+2, column=col, value=value)
        self.workbook.save(self.cfile_name)

#حذف کردن مشتریان    
    def delete_customers(self, row_num):
        self.sheet.delete_rows(row_num+2)
        self.workbook.save(self.cfile_name)
    
#باشگاه مشتریان 

def calculate_discount_credit(customer_id, purchase_amount):
    customers_data = pandas.read_excel('customers.xlsx')
    customer_row = customers_data.loc[customers_data['username'] == customer_id]
    if customer_id in customers_data['username'].values:
        newdiscount = (purchase_amount// 100000) * 5
        newcredit = (purchase_amount// 500000) * 50000
        previous_discount = customer_row['discount'].values[0]
        previous_credit = customer_row['credit'].values[0]
        choice = input("DISCOUNT=d OR CREDIT=c ?")
    
        if choice == 'd':
            if(previous_discount<30):
                payable_amount = purchase_amount *(100-previous_discount)/100
                print(f"PAYABLE AMOUNT: {payable_amount} TOMAN")
            else:
                print("YOUR PREVIOUS DISCOUNT IS MORE THAN 30 ")
            customers_data.loc[customers_data['username'] == customer_id, 'discount'] = newdiscount
            customers_data.to_excel('customers.xlsx', index=False)
        
        elif choice == 'c':
            debt= customer_row['debt'].values[0]
            if debt == 0 :
                    payable_amount = purchase_amount - previous_credit
                    print(f"PAYABLE AMOUNT: {payable_amount} TOMAN")
            
                    customers_data.loc[customers_data['username'] ==customer_id, 'credit'] = newcredit
                    customers_data.to_excel('customers.xlsx', index=False)
            else:
                print("CUSTOMER HAVE DEBT!")
        else:   
            print("WRONG CHOICE")
    else:
            print("USER NOT FOUND!")        
    
        
#فروشگاه های زیر مجموعه
class Shopes:
    def __init__(self, sfile_name):
        self.sfile_name = sfile_name
        self.workbook = openpyxl.load_workbook(sfile_name)
        self.sheet = self.workbook.active


#نمایش دادن فروشگاه های زیر مجموعه 
    def show_shopes(self):
        shopes_list=pandas.read_excel(self.sfile_name)
        print (shopes_list)

#اضافه کردن فروشگاه های زیر مجموعه
    def add_shopes(self, shopes_info):
        self.sheet.append(shopes_info)
        self.workbook.save(self.sfile_name)
        
#تغییر دادن فروشگاه های زیر مجموعه
    def edit_shopes(self, row_num, new_info):
        for col, value in enumerate(new_info, start=1):
            self.sheet.cell(row=row_num+2, column=col, value=value)
        self.workbook.save(self.sfile_name)

#حذف کردن فروشگاه های زیر مجموعه    
    def delete_shopes(self, row_num):
        self.sheet.delete_rows(row_num+2)
        self.workbook.save(self.sfile_name)
   
        
#سفارشات
class Orders:
    def __init__(self, ofile_name):
        self.ofile_file_name = ofile_name
        self.workbook = openpyxl.load_workbook(ofile_name)
        self.sheet = self.workbook.active

# ذخیره کردن سفارشات
    def save_orders(self,order_code, username, shop_code, product, quantity, order_date, delivery_date):
        self.sheet.append(["order Code", "username", "shop code", "food", "quantity", "order Date", "delivery Date"])
        self.sheet.append([ order_code,username, shop_code, product, quantity, order_date, delivery_date])
        self.workbook.save(self.ofile_file_name)
       
def update_order_status(order_code):
    df = pandas.read_excel('orders.xlsx')
    df.loc[df['Order Code'] == order_code, 'Status'] = 'Delivered'
    with pandas.ExcelWriter('orders.xlsx', mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

#تحلیل فروش در پایان هر ماه 
def analyze_monthly_sales(ofile_name):
    data = pandas.read_excel(ofile_name)
    
    low_sales_items = data[data['Total Quantity'] < 10]
    if not low_sales_items.empty:
        print("ITEMS WITH LESS THAN 10 SALES IN THIS MONTH :")
        print(low_sales_items)
    else:
        print("NO ITEMS WITH LESS THAN 10 SALES IN THIS MONTH.")
        print(low_sales_items)

        
#انتخاب گزینه ها 
def main():
    file_name = 'products.xlsx'
    cfile_name='customers.xlsx'
    sfile_name='shopes.xlsx'
    ofile_name='orders.xlsx'
    
    food_products= Food_Products(file_name)
    customers= Customer(cfile_name)
    
    shopes=Shopes(sfile_name)
    orders=Orders(ofile_name)
   
    while True:
        print("1.SHOW PRODUCTS")
        print("2.ADD PRODUCTS")
        print("3.EDIT PRODUCTS")
        print("4.DELETE PRODUCTS")
        print("5.SHOW CUSTOMERS")
        print("6.ADD CUSTOMERS")
        print("7.EDIT CUSTOMERS")
        print("8.DELETE CUSTOMERS")
        print("9.SHOW SHOPES")
        print("10.ADD SHOPES")
        print("11.EDIT SHOPES")
        print("12.DELETE SHOPES")
        print("13.ADD ORDERS")
        print("14.DELIVERY")
        print("15.ANALYZE MONTHLY SALES")
        print("16.CUSTOMERS CLUB")
        print("17.EXIT")

        choice = int(input("ENTER YOUR CHOICE: "))

#محصولات غذایی
        if choice == 1:
            print("HERE OUR PRODUCTS:")
            food_products.show_products()                   
                
        elif choice == 2:
            product_code= int(input("ENTER PRODUCT CODE: "))
            product_name = input("ENTER PRODUCT NAME: ")
            company = input("ENTER COMPANY: ")
            date_of_manufacturing =input ("ENTER DATE OF MAUFACTURING (%Y-%m-%d):")
            expiry = input("ENTER EXPIRY: ")
            bought= int(input("ENTER NO.BOUGHT PRODUCT : "))
            stock= int(input("ENTER NO.STOCK : ")) 
            price= int(input("ENTER PRICE(TOMAN): "))
            print("PRODUCT SUCCESSFULY ADDED.")
            food_products.add_product(product_code, product_name, company, date_of_manufacturing,expiry,bought,stock,price)
       
        elif choice == 3:
            row_num = int(input("ENTER THE ROW NUMBER TO EDIT: "))
            new_info = input(" ENTER CODE:,ENTER NAME:,ENTER COMPANY:,ENTER DATE OF MANUFACTURING:,ENTER EXPIRARY:,ENTER NUMBER OF BOUGHT:,ENTER NUMBER OF STOCK: ,ENTER PRICE: ").split(",")
            print("PRODUCT SUCCESSFULY EDITED.")
            food_products.edit_product(row_num, new_info)
        
        elif choice == 4:
            row_num = int(input("ENTER THE ROW NUMBER TO DELETE: "))
            print("PRODUCT SUCCESSFULY DILITED.")
            food_products.delete_product(row_num)
        
#مشتریان
        elif choice == 5:
            print("HERE OUR CUSTOMERS:")
            customers.show_customers()
        
        elif choice == 6:
            username= input("ENTER USERNAME: ")
            name = input("ENTER CUSTOMER NAME: ")
            family = input("ENTER CUSTOMER FAMILY: ")
            personal_id= int(input("ENTER CUSTOMER PERSONAL ID : "))
            phone_number= input("ENTER CUSTOMER PHONE NUMBER : ") 
            purchase= int(input("ENTER PURCHASE(TOMAN):"))
            discount= int(input("ENTER DISCOUNT:"))
            credit= int(input("ENTER CREDIT(TOMAN):"))
            debt= int(input("ENTER DEBT(TOMAN):"))
            print("CUSTOMER SUCCESSFULY ADDED.")
            customers.add_customers( username,name,family,personal_id,phone_number,purchase,discount,credit,debt)                   
        
        elif choice == 7:
            row_num = int(input("ENTER THE ROW NUMBER TO EDIT: "))
            new_info = input("ENTER USER NAME:,NAME:,FAMILY:,PERSONAL ID:,PHONE NUMBER:,PURCHASE:,DISCOUNT: ,DEBIT: ").split(",")
            print("CUSTOMER SUCCESSFULY EDITED.")
            customers.edit_customers(row_num, new_info)
        
        elif choice == 8:
            row_num = int(input("ENTER THE ROW NUMBER TO DELETE: "))
            print("CUSTOMER SUCCESSFULY DILITED.")
            customers.delete_customers(row_num)
               
#فروشگاه های زیر مجموعه
        if choice == 9:
            print("HERE OUR SHOPES:")
            shopes.show_shopes()
       
        elif choice == 10:
            shopes_info = input("ENTER SHOPE ID:, SHOPE NAME:, NUMBER OF UNDELIVEREED ORDERS:, SELLING PRICE: ").split(",")
            print("SHOPE SUCCESSFULY ADDED.")
            shopes.add_shopes(shopes_info)                   
                
        elif choice == 11:
            row_num = int(input("ENTER THE ROW NUMBER TO EDIT: "))
            new_info = input("ENTER SHOPE ID:, SHOPE NAME:, NUMBER OF UNDELIVEREED ORDERS:, SELLING PRICE: ").split(",")
            print("SHOPE SUCCESSFULY EDITED.")
            shopes.edit_shopes(row_num, new_info)
        
        elif choice == 12:
            row_num = int(input("ENTER THE ROW NUMBER TO DELETE: "))
            print("SHOPE SUCCESSFULY DILITED.")
            shopes.delete_shopes(row_num)    
           
#سفارشات       
        elif choice == 13:
            order_code =int( input("ENTER ORDER CODE: "))
            username = input("ENTER USERNAME: ")
            shop_code = int(input("ENTER SHOPE CODE: "))
            product = input("ENTER PRODUCT: ")
            quantity = int(input("ENTER QUANTITY: "))
            order_date = datetime.now().strftime("%Y-%m-%d")
            delivery_date = input("ENTER DELIVERY DATE: ")
            orders.save_orders(order_code,username, shop_code,product, quantity, order_date, delivery_date)
   
#ثبت تحویل سفارش
        elif choice == 14:
            order_code = input("ENTER THE ORDER CODE: ")
            update_order_status(order_code)

#تحلیل فروش در پایان هر ماه 
        elif choice == 15:
            analyze_monthly_sales(ofile_name)
   
#باشگاه مشتریان  
        elif choice==16:
            customer_id = input("ENTER USER NAME : ")
            purchase_amount = int(input("PLEASE ENTER PURCHASE (TOMAN) :" ))
            calculate_discount_credit(customer_id,purchase_amount)

#خروج
        elif choice == 17:
            break
        
# انتخاب اشتباه     
        else:
            print("WRONG CHOICE!")

if __name__ == "__main__":
    main()
